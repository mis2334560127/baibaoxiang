"""
百宝箱 - 批量合并 Excel 模块

功能：
  1. 支持直接添加 .xlsx / .xls / .xlsm 文件
  2. 也支持从压缩包（.zip / .rar / .7z）中提取 Excel 文件
  3. 读取所有 Excel 数据（假设格式一致），合并写入到单个 Excel 文件
"""
import os
import zipfile
import tempfile
import shutil
from pathlib import Path

# 支持的压缩包格式 → 扩展名集合
ARCHIVE_EXTS = {'.zip', '.rar', '.7z'}

# Excel 扩展名（直接支持的 Excel 文件格式 + 压缩包内提取的格式）
EXCEL_EXTS = {'.xlsx', '.xls', '.xlsm'}

# 所有支持的文件扩展名（压缩包 + 直接 Excel 文件）
ALL_SUPPORTED_EXTS = ARCHIVE_EXTS | EXCEL_EXTS


def extract_excel_from_archive(archive_path: str, temp_dir: str) -> list[str]:
    """
    从单个压缩包中提取所有 Excel 文件到临时目录。
    返回提取后的 Excel 文件绝对路径列表。
    自动过滤掉 __MACOSX 和临时文件（~$开头）。
    """
    ext = os.path.splitext(archive_path)[1].lower()
    extracted: list[str] = []

    if ext == '.zip':
        with zipfile.ZipFile(archive_path, 'r') as zf:
            for name in zf.namelist():
                base = os.path.basename(name)
                if base.startswith('~$') or name.startswith('__MACOSX'):
                    continue
                if os.path.splitext(name)[1].lower() in EXCEL_EXTS:
                    zf.extract(name, temp_dir)
                    extracted.append(os.path.join(temp_dir, name))

    elif ext == '.rar':
        try:
            import rarfile
        except ImportError:
            raise ImportError("缺少 rarfile 库，请执行: pip install rarfile")
        with rarfile.RarFile(archive_path, 'r') as rf:
            for name in rf.namelist():
                base = os.path.basename(name)
                if base.startswith('~$') or name.startswith('__MACOSX'):
                    continue
                if os.path.splitext(name)[1].lower() in EXCEL_EXTS:
                    rf.extract(name, temp_dir)
                    extracted.append(os.path.join(temp_dir, name))

    elif ext == '.7z':
        try:
            import py7zr
        except ImportError:
            raise ImportError("缺少 py7zr 库，请执行: pip install py7zr")
        with py7zr.SevenZipFile(archive_path, 'r') as szf:
            all_names = szf.getnames()
            excel_names = [
                n for n in all_names
                if os.path.splitext(n)[1].lower() in EXCEL_EXTS
                and not os.path.basename(n).startswith('~$')
                and not n.startswith('__MACOSX')
            ]
            if excel_names:
                szf.extract(path=temp_dir, targets=excel_names)
                extracted = [os.path.join(temp_dir, n) for n in excel_names]

    return extracted


def read_excel_data(file_path: str) -> tuple[list, list[list]]:
    """
    读取单个 Excel 文件的数据。

    Returns:
        (headers, rows)
        - headers: 第一行作为表头列表
        - rows: 数据行列表（每行是一个列表），自动跳过完全空行
    """
    ext = os.path.splitext(file_path)[1].lower()

    if ext in ('.xlsx', '.xlsm'):
        from openpyxl import load_workbook
        wb = load_workbook(file_path, read_only=True, data_only=True)
        ws = wb.active
        headers: list = []
        rows: list[list] = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            row_data = list(row)
            if i == 0:
                headers = [v if v is not None else '' for v in row_data]
            else:
                if any(v is not None and str(v).strip() != '' for v in row_data):
                    rows.append(row_data)
        wb.close()
        return headers, rows

    elif ext == '.xls':
        try:
            import xlrd
        except ImportError:
            raise ImportError("缺少 xlrd 库，请执行: pip install xlrd")
        wb = xlrd.open_workbook(file_path)
        ws = wb.sheet_by_index(0)
        headers = [ws.cell_value(0, c) for c in range(ws.ncols)]
        rows = []
        for r in range(1, ws.nrows):
            row_data = [ws.cell_value(r, c) for c in range(ws.ncols)]
            if any(str(v).strip() != '' for v in row_data):
                rows.append(row_data)
        return headers, rows

    return [], []


def merge_and_write(headers: list, all_rows: list[list], output_path: str):
    """
    将合并后的表头和数据行写入单个 .xlsx 文件。
    """
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    if ws is None:
        ws = wb.create_sheet()

    # 写表头
    for c, header in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=header)

    # 写数据
    for r, row in enumerate(all_rows, 2):
        for c, value in enumerate(row, 1):
            ws.cell(row=r, column=c, value=value)

    wb.save(output_path)
